"""
Easify topical spreadsheet - row appending
==========================================

Adds one row per question to `easify-topical-template.xlsx` without disturbing
anything already in it.

How it stays safe:

  * your original template is never written to. The first time you add a row
    it is copied into the output folder, and every later row goes into that
    copy - so the master stays pristine and the copy is what you hand over.
  * columns are found by their heading, not by position, so a sheet whose
    columns sit somewhere else still works.
  * only columns A-I are written. Columns J and K fill themselves, and the
    brief says not to touch them - so the formulas from the row above are
    copied down instead, which is what Excel would do if you dragged the row.
  * saving a question and then its mark scheme updates the *same* row rather
    than making a second one.
"""

from __future__ import annotations

import re
import shutil
from copy import copy
from dataclasses import dataclass
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - startup guard for contributors
    raise ImportError(
        "openpyxl is missing - that is what writes the spreadsheet.\n"
        "Run this once, then try again:\n\n"
        "    python -m pip install -r requirements.txt\n"
    ) from None

#: Columns A-I of the Questions tab - the ones a contributor fills in.
FIELDS = ("subject", "paper", "chapter", "subtopic", "year", "difficulty",
          "marks", "q_filename", "ms_filename", "youtube_url")

#: Spellings that have been seen for each heading, so a slightly different
#: template still lines up.
ALIASES: dict[str, tuple[str, ...]] = {
    "subject": ("subject",),
    "paper": ("paper", "paperno", "papernumber"),
    "chapter": ("chapter", "topic", "chaptername"),
    "subtopic": ("subtopic", "subtopics", "subchapter", "subchapters",
                 "subtopicname", "section"),
    "year": ("year", "examyear"),
    "difficulty": ("difficulty", "level"),
    "marks": ("marks", "totalmarks", "mark"),
    "q_filename": ("qfilename", "questionfilename", "qfile", "questionfile"),
    "ms_filename": ("msfilename", "markschemefilename", "msfile", "markschemefile"),
    "youtube_url": ("youtubeurl", "youtube", "video", "videourl", "solutionvideo"),
}

SHEET_NAMES = ("questions", "question")


class SheetError(RuntimeError):
    """Something about the workbook stopped us - always worth showing the user."""


@dataclass
class AddResult:
    path: Path
    row: int
    updated: bool          # True = filled in an existing question's row
    sheet: str


def _norm(text: object) -> str:
    """'Q Filename' and 'q_filename' both become 'qfilename'."""
    return re.sub(r"[^a-z0-9]", "", str(text).lower())


# --------------------------------------------------------------------------
# Finding our way around the workbook
# --------------------------------------------------------------------------

def find_sheet(workbook):
    """The Questions tab, or the only sensible alternative."""
    for name in workbook.sheetnames:
        if _norm(name) in SHEET_NAMES:
            return workbook[name]
    for sheet in workbook.worksheets:            # fall back to the first tab
        return sheet
    raise SheetError("That workbook has no sheets in it.")


def find_header(sheet) -> tuple[int, dict[str, int]]:
    """
    Locate the heading row and map each field to its column number.

    Searched by name across the first 15 rows, so the headings do not have to
    start in row 1 or sit in the brief's order.
    """
    lookup = {alias: field for field, names in ALIASES.items() for alias in names}
    lookup.update({field: field for field in FIELDS})

    best: tuple[int, dict[str, int]] = (0, {})
    for row in range(1, min(sheet.max_row, 15) + 1):
        found: dict[str, int] = {}
        for column in range(1, min(sheet.max_column, 40) + 1):
            field = lookup.get(_norm(sheet.cell(row=row, column=column).value))
            if field and field not in found:
                found[field] = column
        if len(found) > len(best[1]):
            best = (row, found)

    row, mapping = best
    if "q_filename" not in mapping or "subject" not in mapping:
        raise SheetError(
            "Could not find the column headings in that spreadsheet.\n\n"
            "The tool looks for a row containing 'subject' and 'q_filename'.\n"
            "Open the file and check you picked the right one - it should be "
            "easify-topical-template.xlsx with a Questions tab."
        )
    return row, mapping


def last_data_row(sheet, header_row: int, mapping: dict[str, int]) -> int:
    """
    The last row holding real data.

    Only columns A-I are consulted. The auto columns often carry formulas
    hundreds of rows further down, and counting those would leave a huge gap
    of empty rows before the new one.
    """
    columns = sorted(mapping.values())
    last = header_row
    # iter_rows in bulk: the real template carries formulas down to row 6000,
    # so this is thousands of cells and reading them one at a time is slow.
    for offset, values in enumerate(sheet.iter_rows(
            min_row=header_row + 1, max_col=columns[-1], values_only=True)):
        if any(values[c - 1] not in (None, "") for c in columns):
            last = header_row + 1 + offset
    return last


def _stem(filename: object) -> str:
    """'physics_p11_dynamics_2025mj_q19_Q.png' -> the part before _Q / _MS."""
    text = str(filename or "").strip()
    if not text:
        return ""
    return re.sub(r"_(Q|MS)\.png$", "", text, flags=re.IGNORECASE)


def find_question_row(sheet, header_row: int, mapping: dict[str, int],
                      stem: str) -> int | None:
    """The row this question already occupies, if it has one."""
    if not stem:
        return None
    columns = [mapping[f] for f in ("q_filename", "ms_filename") if f in mapping]
    for row in range(header_row + 1, last_data_row(sheet, header_row, mapping) + 1):
        for column in columns:
            if _stem(sheet.cell(row=row, column=column).value).lower() == stem.lower():
                return row
    return None


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------

def _carry_down_extras(sheet, mapping: dict[str, int], row: int) -> None:
    """
    Copy the self-filling columns (J, K, anything else we do not own) down
    from the row above, translating their cell references - exactly what Excel
    does when you drag a row down.
    """
    if row < 2:
        return
    ours = set(mapping.values())
    for column in range(1, sheet.max_column + 1):
        if column in ours:
            continue
        target = sheet.cell(row=row, column=column)
        if target.value not in (None, ""):
            continue                                    # already filled in
        above = sheet.cell(row=row - 1, column=column)
        if isinstance(above.value, str) and above.value.startswith("="):
            origin = f"{get_column_letter(column)}{row - 1}"
            target.value = Translator(above.value, origin=origin).translate_formula(
                f"{get_column_letter(column)}{row}")


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    """Keep the new row looking like the rows above it."""
    if source_row < 1 or source_row == target_row:
        return
    for column in range(1, sheet.max_column + 1):
        source = sheet.cell(row=source_row, column=column)
        sheet.cell(row=target_row, column=column)._style = copy(source._style)


def _extend_tables(sheet, last_row: int) -> None:
    """Grow any Excel table so the new row is inside it (keeps auto-fill alive)."""
    for table in getattr(sheet, "tables", {}).values():
        try:
            start, end = str(table.ref).split(":")
            end_col = re.match(r"([A-Z]+)(\d+)", end)
            start_row = int(re.match(r"([A-Z]+)(\d+)", start).group(2))
            if end_col and start_row <= last_row and int(end_col.group(2)) < last_row:
                table.ref = f"{start}:{end_col.group(1)}{last_row}"
        except (AttributeError, ValueError):
            continue                                    # not worth failing over


def values_for(meta, filename: str = "") -> dict[str, object]:
    """
    Turn the form's contents into cell values, following the brief.

    The paper column holds the paper number only - the variant belongs in the
    file name - and the chapter is written in its slug form so it is spelled
    identically on every row.

    Both file name columns are filled from the one entry. The two names differ
    only by their _Q / _MS ending, so there is nothing to gain from typing the
    mark scheme in separately - and plenty to lose, since that is exactly
    where a typo would break the link between sheet and image.
    """
    from photofuse import slugify, tidy_subtopics

    values: dict[str, object] = {
        "subject": slugify(meta.subject),
        # Written the way a student reads it, not slugified. The brief is
        # explicit that the squashed spelling belongs in the file name while
        # the sheet's chapter column keeps its spaces and capitals - that
        # column is what the site filters by and shows.
        "chapter": " ".join(str(meta.chapter).split()),
        "subtopic": tidy_subtopics(getattr(meta, "subtopic", "")),
        "difficulty": str(meta.difficulty).strip(),
        "youtube_url": str(meta.youtube_url).strip(),
        "q_filename": meta.filename("Q"),
        "ms_filename": meta.filename("MS"),
    }

    paper = str(meta.paper).strip()
    values["paper"] = int(paper) if paper.isdigit() else paper
    year = str(meta.year).strip()
    values["year"] = int(year) if year.isdigit() else year
    marks = str(meta.marks).strip()
    values["marks"] = int(marks) if marks.isdigit() else marks

    return values


def add_row(workbook_path: str | Path, meta) -> AddResult:
    """
    Put this question into the sheet, updating its row if it already has one.

    Both the question and the mark scheme file names are written from the one
    entry, so you only ever fill the form in once per question.
    """
    path = Path(workbook_path)
    if not path.is_file():
        raise SheetError(f"No spreadsheet at:\n{path}")

    try:
        workbook = load_workbook(path)           # keep formulas, not their results
    except PermissionError:
        raise SheetError(f"{path.name} is open in Excel. Close it and try again.") from None
    except Exception as exc:                     # noqa: BLE001 - surfaced to the user
        raise SheetError(f"Could not open {path.name}: {exc}") from None

    sheet = find_sheet(workbook)
    header_row, mapping = find_header(sheet)

    existing = find_question_row(sheet, header_row, mapping, meta.stem)
    row = existing or last_data_row(sheet, header_row, mapping) + 1
    if row <= header_row:
        row = header_row + 1

    if existing is None:
        _copy_row_style(sheet, row - 1, row)

    for field, value in values_for(meta).items():
        column = mapping.get(field)
        # An empty box is never written, so a field left blank can neither
        # wipe what is already in the cell nor add clutter to a fresh row.
        # Anything the sheet has no column for is simply skipped.
        if column is None or value in (None, ""):
            continue
        sheet.cell(row=row, column=column).value = value

    _carry_down_extras(sheet, mapping, row)
    _extend_tables(sheet, row)

    try:
        workbook.save(path)
    except PermissionError:
        raise SheetError(f"{path.name} is open in Excel. Close it and try again.") from None
    finally:
        workbook.close()

    return AddResult(path=path, row=row, updated=existing is not None, sheet=sheet.title)


def working_copy(template: str | Path, out_dir: str | Path) -> Path:
    """
    The copy of the spreadsheet that rows are actually added to.

    Made once, inside the output folder, so the template you were given is
    never modified. Pick that copy next time and rows keep stacking up in it.
    """
    template = Path(template)
    out_dir = Path(out_dir)
    if not template.is_file():
        raise SheetError(f"No spreadsheet at:\n{template}")

    out_dir.mkdir(parents=True, exist_ok=True)
    destination = out_dir / template.name
    if destination.resolve() == template.resolve():
        return template                           # already working in the copy
    if not destination.exists():
        shutil.copy2(template, destination)
    return destination
